import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from openpyxl.utils import get_column_letter
import platform
import subprocess
import time

# --- funcție helper pentru WSL ---
def wsl_to_windows_path(path):
    if path.startswith("/mnt/"):
        drive_letter = path[5]  # litera discului
        win_path = f"{drive_letter.upper()}:\\{path[7:].replace('/', '\\')}"
        return win_path
    return path

def create_excel_template(file_path=None):
    if file_path is None:
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        PROJECT_DIR = os.path.join(BASE_DIR, "..")
        OUTPUT_DIR = os.path.join(PROJECT_DIR, "data", "input_samples")
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        file_name = f"Facturi_Template_{timestamp}.xlsx"
        file_path = os.path.join(OUTPUT_DIR, file_name)

    columns = [
        "Număr factură", "Data emiterii", "Tip factură", "Monedă",
        "Nume cumpărător", "ID legal cumpărător", "ID TVA cumpărător",
        "Stradă cumpărător", "Oraș cumpărător", "Județ cumpărător",
        "Cod poștal cumpărător", "Țară cumpărător",
        "Termeni plată", "Linii factură (produse)",
        "Valoare totală fără TVA", "Total TVA", "Total plată"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturi"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.protection = Protection(locked=True)

    for row in ws.iter_rows(min_row=2, max_row=100, max_col=len(columns)):
        for cell in row:
            cell.protection = Protection(locked=False)

    for i, col_name in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(len(col_name) + 2, 15)

    ws.protection.sheet = True
    ws.protection.password = "invoice"

    wb.save(file_path)
    print(f"✓ Excel template creat: {file_path}")

    # Deschide fișierul automat
    system_platform = platform.system()
    try:
        if system_platform == "Windows":
            os.startfile(file_path)
        elif system_platform == "Darwin":  # macOS
            subprocess.call(["open", file_path])
        else:  # Linux
            subprocess.call(['cmd.exe', '/c', 'start', '', wsl_to_windows_path(file_path)])
    except Exception as e:
        print(f"⚠️ Nu s-a putut deschide Excel-ul automat: {e}")
        
    return file_path